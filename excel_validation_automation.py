import string
import time
import os
import openpyxl as excel
from openpyxl.styles import PatternFill, Font
import pandas as pd
from openpyxl.utils.datetime import to_excel

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
# --------------------------------------------------------------------------------------------------
red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Red
orange = PatternFill(start_color="FFFF9900", end_color="FFFF9900", fill_type="solid")  # Orange
yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Yellow
blue = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")  # Blue
green = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")  # Green
# --------------------------------------------------------------------------------------------------

print("\nApplication Starting...")
time.sleep(3)
print("\nLoading...")
time.sleep(3)

task_ia = """
>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> Task IA <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

Explaining Task Impact Analysis.
----------------------------------------------------------------------------------------------------------------"""
print(task_ia)
time.sleep(2)

manual_guide = """
>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>> Task Steps <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

----------------------------------------------------------------------------------------------------------------
Explaining Task Steps.

----------------------------------------------------------------------------------------------------------------
"""
print(manual_guide)
time.sleep(4)

print("\nReading File....")
time.sleep(3)

# -------------------
# Read Sheet.
# -------------------
main_workbook_path = f"{os.getcwd()}\\excel file name.xlsx"
# -------------------

path_lives = 0

while path_lives != 3  :
    print("\nChecking The File Accessibility.....")
    time.sleep(2)

    if os.path.exists(main_workbook_path):
        print("\nExcel file exists.")
        time.sleep(2)
        break

    else:
        path_lives += 1
        print("\nExcel file ( excel file name ) doesn't exist.")
        input(f"\nAdd the file and press Enter to check again.")
        continue

if path_lives == 3 :
    print("\nAll Tries Consumed, please try again later")
    exit()
# --------------------------------------------------------------------------------------------------

# Read File
pandas_eye = pd.read_excel( main_workbook_path )

# Remove empty column border.
pandas_eye = pandas_eye.iloc[ : , 1: ]

# Remove empty row border and select header names.
pandas_eye.columns = pandas_eye.iloc[0]

pure_df = pandas_eye.copy()

# ----------------------------------------------------------------------------------------------
# data cleaning---------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------
print("\nCleaning in progress......")

pure_df = pure_df[pure_df["Account Number"].str.startswith("1")]

pure_df = pure_df[ ~pure_df["Status"].isin(["Suspended","Deactive"]) ]

pure_df = pure_df.dropna(subset=["Msisdn"])

pure_df = pure_df[   ~pure_df["Rate Plan Desc"].str.contains(
    "Rate Plan Name"  , na=False)    ]

pure_df = pure_df[ ~pure_df["First Name"].str.contains("Test Name Convension")   ]

time.sleep(5)
# ----------------------------------------------------------------------------------------------
# divide, distribute and save-------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------
print("\nDistribute sheets.....")
time.sleep(2)
print("\nLoading..")
time.sleep(2)
print("\nLoading......\n")
time.sleep(3)


# ----------------------------------------------------------------------------------------------
# 1 >> Dummy.
print("\nExtract Dummy Sheet ")
dummy_df = pure_df[
    pure_df["First Name"].str.contains("names|names") |
    pure_df["Last Name"].str.contains("names|names")
]
dummy_df.to_excel(f"{os.getcwd()}\\excel name.xlsx" , index=False)
time.sleep(2)
print("\nDummy Sheet Saved")
time.sleep(2)
# ----------------------------------------------------------------------------------------------
print("""
----------------------------------------------------------------------------------------------------------------
For Dummy Sheet please note :

Explaining dummy task steps.

----------------------------------------------------------------------------------------------------------------""")
input("\nPress Enter to continue, After Reading the above well, (Don't open any excel sheet now till finish).\n")

# ----------------------------------------------------------------------------------------------
# 2 >> E-gated.
print("-"*50)
print("\nExtract E-Gated Sheet")
Gated_df = pure_df[ pure_df["Rate Plan Desc"].str.contains("Gated")  ]

Gated_df.to_excel(f"{os.getcwd()}\\E-Gated name sheet.xlsx" , index=False)
time.sleep(2)
print("\nE-Gated Sheet Saved")
print("-"*50)
# ----------------------------------------------------------------------------------------------

# ----------------------------------------------------------------------------------------------
# 3 >> Full English Sheet.
print("-"*50)
print("\nExtract English Sheet")
english_df = pure_df[ pure_df["First Name"].str.contains("a|b|c|d|e|f|g|h|i|j|k|l|m|n|o|p|q|r|s|t|u|v|w|x|y|z|A|B|C|D|E|F|G|H|I|J|K|L|M|N|O|P|Q|R|S|T|U|V|W|X|Y|Z") ]
english_df = english_df[~english_df["Rate Plan Desc"].str.contains("Gated")  ]
english_df = english_df[~english_df["First Name"].str.contains("Dummy|dummy") ]

english_df.to_excel(f"{os.getcwd()}\\English Sheet.xlsx" , index=False)
time.sleep(2)
print("\nEnglish Sheet Saved")
print("-"*50)
# ----------------------------------------------------------------------------------------------

print("\nSaving Main Sheet")

pure_df = pure_df[~pure_df["First Name"].str.contains("Dummy|dummy") ]

pure_df = pure_df[~pure_df["Rate Plan Desc"].str.contains("Gated")  ]

pure_df = pure_df[~pure_df["First Name"].str.contains("a|b|c|d|e|f|g|h|i|j|k|l|m|n|o|p|q|r|s|t|u|v|w|x|y|z|A|B|C|D|E|F|G|H|I|J|K|L|M|N|O|P|Q|R|S|T|U|V|W|X|Y|Z")  ]

pure_df.to_excel(main_workbook_path , index=False , sheet_name="OnDemand Bad Bills")

time.sleep(3)

print("\nMain Sheet Saved")

# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------
time.sleep(2)
print("\nHandling the file path.")

# Create Untitled Sheet.
untitled_path = f"{os.getcwd()}\\untitled Sheet.xlsx"

untitled_df =  pure_df[ pure_df["Title"].isna() ]

untitled_df.to_excel( untitled_path , index=False , sheet_name="untitled")

# ----------------------------------------------------------------------------------------------
# openpyxl handling.
# ----------------------------------------------------------------------------------------------

# Calling main sheet.
main_wb_1 = excel.load_workbook( main_workbook_path )
main_sh_1 = main_wb_1["sheet name"]

# ------------------------------------------

# Calling untitled sheet.
untitled_wb_2 = excel.load_workbook( untitled_path )
untitled_sh_2 = untitled_wb_2["sheet name"]

# ----------------------------------------------------------------------------------------------
# functions zone.
time.sleep(3)
print("\nStarting the functionality....")
time.sleep(1)
print("\nLoading.......")
time.sleep(2)
print("\nStill Loading...........")
time.sleep(4)

# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------


def corrupted_names(param_sheet_1,col_seq_1) :
    container = ["names not preferred" , 
                 *string.punctuation, *string.digits, *string.ascii_letters]
    # ----------------------------------------------
    def check(dynamic_cell) :
        if dynamic_cell in container:
            return 0

        else:
            for x in dynamic_cell:
                if x in container:
                    return 0
    # ----------------------------------------------

    for name_cell in range( 2 , param_sheet_1.max_row + 1 ) :

        cell_step = param_sheet_1.cell(name_cell, col_seq_1)

        if cell_step.value is None :
            cell_step.fill = red

        elif isinstance(cell_step.value, int) :
            cell_step.fill = red


        elif isinstance(cell_step.value, str) :

            if len(str(cell_step.value)) <= 1:
                cell_step.fill = red

            else :
                check_fx = check(cell_step.value)
                if check_fx == 0 :
                    cell_step.fill = red

        else :
            cell_step.fill = yellow


# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------


def address_issues(param_sheet_2,col_seq_2) :
    add = r"\\"
    punctuation_without_hyphen = [char for char in string.punctuation if char not in {'-', '/', ',', '.','=','*','_',add}]

    container = ["names not preferred",
                 *punctuation_without_hyphen, "?"]
    # ------------------------------------------------------------------------------------------------------------------
    def check(dynamic_cell):
        if dynamic_cell in container :
            return 0

        else :
            for x in dynamic_cell:
                if x in container:
                    return 0
    # -------------------------------------------------------------------
    for address_cell in range( 2, param_sheet_2.max_row + 1 ) :

        cell_step = param_sheet_2.cell( address_cell, col_seq_2 )

        if cell_step.value is None:
            cell_step.fill = red

        elif isinstance(cell_step.value, int):
            cell_step.fill = red

        elif "?" in cell_step.value:
            cell_step.fill = red

        elif isinstance(cell_step.value, str):

            if len(str(cell_step.value)) <= 1:
                cell_step.fill = red

            else :
                check_fx = check(cell_step.value)
                if check_fx == 0 :
                    cell_step.fill = red

        else:
            cell_step.fill = red


# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------


# ----------------------------------------------------------------------------------------------
print("\nHandling corrupted names issue.")
for col_num in [6,7] :
    corrupted_names( main_sh_1, col_num )

time.sleep(4)
print("\nDone.")
# ----------------------------------------------------

print("\nHandling Address issue.")
address_issues( main_sh_1 , 8 )

time.sleep(4)
print("\nDone.")
# ----------------------------------------------------

time.sleep(2)
print("\nSaving The Sheet.")
time.sleep(2)
main_wb_1.save( main_workbook_path )
print("\nSheet Saved.")
time.sleep(2)

# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------

print("\nStarting the handling of untitled sheet.")

names_array_path = f"{os.getcwd()}\\Names-Array.xlsx"

names_book = excel.load_workbook( names_array_path )

names_sheet_males = names_book["males"]

names_sheet_females = names_book["females"]

# ----------------------------------------------------------------------------------------------
print("\nLoading...")

males_array = []

for male in range( 2 , names_sheet_males.max_row + 1 ) :

    cell_males = names_sheet_males.cell( male , 1 )

    males_array.append( cell_males.value )

# ------------------------------------

fe_males_array = []

for female in range( 2, names_sheet_females.max_row + 1 ) :

    cell_females = names_sheet_females.cell( female , 1 )

    fe_males_array.append( cell_females.value )

# ----------------------------------------------------------------------------------------------
print("\nLoading......")

def untitled(param_sheet_3) :

    for per_name in range( 2 , param_sheet_3.max_row + 1 ) :

        cell_name = param_sheet_3.cell( per_name , 6 )

        if cell_name.value is None :
             cell_name.fill = blue

        elif cell_name.value in males_array :
            cell_name.fill = orange

        elif cell_name.value in fe_males_array :
             cell_name.fill = green

        else:
             cell_name.fill = yellow


# ----------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------
print("\nSaving Data...")

untitled( untitled_sh_2 )

time.sleep(2)

untitled_wb_2.save( untitled_path )

print("\nData Saved.")

time.sleep(2)

input("\nPress enter to exit.")

time.sleep(2)

input("\nPress enter again to exit.")

time.sleep(2)

print("\nExiting...")

time.sleep(2)

print("\nDone\n")

time.sleep(3)